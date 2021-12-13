from django.db import models
from django.utils import timezone
from django.shortcuts import reverse
from django.contrib.auth.models import User
from django.dispatch import receiver
from django.db.models.signals import post_save
from slugify import slugify
from django.utils.translation import gettext_lazy as _
from simple_history.models import HistoricalRecords
from mptt.models import MPTTModel, TreeForeignKey
from PIL import Image
from datetime import date
import os
import calendar


def get_base_components(all_parts):
    base_components = {}

    for part in all_parts:
        simple_components = part.base.simple_components.filter(simple_object__base_object__isnull=False)
        if simple_components:
            for simple_component in simple_components:
                base_object = simple_component.simple_object.base_object
                if len(base_object.simpleobject_set.all()) == 1:
                    if base_object in base_components:
                        base_components[base_object] = base_components[base_object] + simple_component.amount
                    else:
                        base_components[base_object] = simple_component.amount
                else:
                    base_components[base_object] = '-'

    if len(base_components) > 0:
        return base_components
    else:
        return None


def get_all_children(big_object, all_children=None):
    if all_children is None:
        all_children = list()
    children = big_object.get_children()
    for child in children:
        all_children.append(child)
        if child.get_children():
            get_all_children(child)

    return all_children


def gen_slug(title, all_slugs, lab=None):
    new_slug = slugify(title, to_lower=True, separator='_')
    if lab is not None:
        new_slug = '{}_{}'.format(lab, new_slug)

    base_slug = new_slug
    end = 1
    while True:
        if new_slug in all_slugs:
            new_slug = '{}_{}'.format(base_slug, end)
            end += 1
        else:
            break

    return new_slug


def gen_slug_for_categories(title):
    new_slug = slugify(title, to_lower=True, separator='_')
    return new_slug


def gen_text_price(int_price):
    object_price = '{:,}'.format(int_price).replace(',', ' ').replace('.', ',')
    if len(object_price.split(',')[1]) == 1:
        object_price += '0'
    new_price = '{},{}'.format(object_price.split(',')[0], object_price.split(',')[1][0:2])
    return new_price


def update_big_objects_price(base_big_object, update_child_price=True):
    """Обновляет стоимость всех сложных объектов верхнего уровня в которых есть данный базовый сложный объект"""
    big_objects = base_big_object.get_top_level_big_objects()
    for big_object in big_objects:
        if big_object.status in ['NW', 'IW', 'RD']:
            big_object.update_total_price(update_child_price=update_child_price)


def some_model_thumb_name(instance, filename):
    original_image_path = str(instance.image).rsplit('/', 1)[0]
    return os.path.join(original_image_path, filename)


def some_model_save_photo(self, width, height):
    extension = str(self.image.path).rsplit('.', 1)[1]  # получаем расширение загруженного файла
    filename = str(self.image.path).rsplit(os.sep, 1)[1].rsplit('.', 1)[0]  # получаем имя загруженного файла (без пути к нему и расширения)
    full_path = str(self.image.path).rsplit(os.sep, 1)[0]  # получаем путь к файлу (без имени и расширения)

    img = Image.open(self.image.path)
    big_img = Image.open(str(self.image.path))
    thumb_name = filename + str("_big") + '.' + extension
    big_img.save(full_path + os.sep + thumb_name, quality=95)
    self.image_big = some_model_thumb_name(self, thumb_name)
    img.save(self.image.path, quality=60)


def generate_path(instance, filename):
    if instance.category.simple_object:
        return '{0}/images/{1}'.format(instance.category.simple_object.slug, filename)
    elif instance.category.big_object:
        return '{0}/images/{1}'.format(instance.category.big_object.slug, filename)
    elif instance.category.invoice:
        return 'invoices/{0}/images/{1}'.format(instance.category.invoice.number, filename)
    else:
        return 'images/{0}'.format(filename)


def generate_path_for_files(instance, filename):
    if instance.category.simple_object:
        return '{0}/files/{1}'.format(instance.category.simple_object.slug, filename)
    elif instance.category.big_object:
        return '{0}/files/{1}'.format(instance.category.big_object.slug, filename)
    elif instance.category.invoice:
        return '{0}/files/{1}'.format(instance.category.invoice.number, filename)
    else:
        return 'files/{0}'.format(filename)


def generate_path_for_database(instance, filename):
    return 'database/{0}/{1}'.format(instance.lab.name, filename)


def generate_path_for_avatar(instance, filename):
    slug_name = slugify(instance.name, to_lower=True, separator='_')
    return 'avatars/{0}/{1}'.format(slug_name, filename)


def del_path(full_path):
    try:
        if len(os.listdir(full_path)) == 0:
            os.rmdir(full_path)
    except Exception as err:
        print(err)
        pass

    full_path = full_path.rsplit(os.sep, 1)[0]
    try:
        if len(os.listdir(full_path)) == 0:
            os.rmdir(full_path)
    except Exception as err:
        print(err)
        pass


class LabName(models.Model):
    name = models.CharField(max_length=200, unique=True, verbose_name='Название лаборатории')
    slug = models.SlugField(max_length=250, unique=True, blank=True, verbose_name='URL')

    class Meta:
        verbose_name = 'Лаборатория'
        verbose_name_plural = 'Лаборатории'
        ordering = ['name']

    def __str__(self):
        return self.name

    def save(self, *args, **kwargs):
        if self.id:
            old_self = LabName.objects.get(pk=self.pk)
            if old_self.name != self.name:
                self.create_slug()
        else:
            self.create_slug()

            # self.slug = gen_slug(lab=False, title=self.name)
            # AllSlugs(slug=self.slug, lab='LB').save()
        super().save(*args, **kwargs)

    def get_absolute_url(self):
        return reverse('categories_list_url', kwargs={'lab': self.slug})

    def lab_main_page_url(self):
        return reverse('lab_main_page', kwargs={'lab': self.slug})

    def create_slug(self):
        all_slugs = LabName.objects.all().values_list('slug', flat=True)
        self.slug = gen_slug(title=self.name, all_slugs=all_slugs)


class Room(models.Model):
    number = models.CharField(verbose_name='Номер/название кабинета', max_length=100, unique=True)
    lab = models.ForeignKey(LabName, on_delete=models.CASCADE, related_name='room',
                            verbose_name='Лаборатория', blank=True, null=True)
    slug = models.SlugField(max_length=250, unique=True, blank=True, verbose_name='URL')

    class Meta:
        verbose_name = 'Кабинет'
        verbose_name_plural = 'Кабинеты'
        ordering = ['number']

    def __str__(self):
        return self.number

    def save(self, *args, **kwargs):
        if self.id:
            old_self = Room.objects.get(pk=self.pk)
            if old_self.number != self.number:
                self.create_slug()
        else:
            self.create_slug()
        super().save(*args, **kwargs)

    def create_slug(self):
        all_slugs = BaseObject.objects.all().values_list('slug', flat=True)
        self.slug = gen_slug(title=self.number, all_slugs=all_slugs)

    def get_absolute_url(self):
        return reverse('room_page_url', kwargs={'slug': self.slug, 'lab': self.lab.slug})


class Category(models.Model):
    class ChoicesObjectCat(models.TextChoices):
        BASE_OBJECT = 'BO', _('Базовый объект')
        SIMPLE_OBJECT = 'SO', _('Простой объект')
        BIG_OBJECT = 'BG', _('Составной объект')

    class ChoicesObjectType(models.TextChoices):
        DEF = 'DF', _('----------')
        EQUIPMENT = 'EQ', _('Оборудование')
        MATERIALS = 'MT', _('Материалы')

    name = models.CharField(max_length=200, verbose_name='Название категории')
    slug = models.SlugField(max_length=250, unique=True, blank=True, verbose_name='URL')
    lab = models.ForeignKey(LabName, on_delete=models.CASCADE, verbose_name='Лаборатория')
    text = models.TextField(blank=True, verbose_name='Описание')
    cat_type = models.CharField(
        max_length=2,
        choices=ChoicesObjectCat.choices,
        default=ChoicesObjectCat.SIMPLE_OBJECT,
        verbose_name='Тип категории'
    )
    obj_type = models.CharField(
        max_length=2,
        choices=ChoicesObjectType.choices,
        default=ChoicesObjectType.DEF,
        verbose_name='Тип объекта'
    )

    class Meta:
        verbose_name = 'Категория'
        verbose_name_plural = 'Категории'
        ordering = ['name']

    def __str__(self):
        return f'{self.lab.name} : {self.name}'

    def save(self, *args, **kwargs):
        if self.pk is not None:
            old_self = Category.objects.get(pk=self.pk)
            if old_self.name != self.name:
                self.create_slug()
        else:
            self.create_slug()
        # self.slug = gen_slug(lab=self.lab.slug, title=self.name)
        # AllSlugs(slug=self.slug, lab=self.lab, cat_type='CT').save()
        super().save(*args, **kwargs)

    def get_absolute_url(self):
        return reverse('category_page_url', kwargs={'slug': self.slug, 'lab': self.lab.slug})

    def create_slug(self):
        all_slugs = Category.objects.all().values_list('slug', flat=True)
        self.slug = gen_slug(lab=self.lab.slug, title=self.name, all_slugs=all_slugs)


class Position(models.Model):
    name = models.CharField(verbose_name='Название должности', max_length=200)

    class Meta:
        verbose_name = 'Должность'
        verbose_name_plural = 'Должности'

    def __str__(self):
        return self.name


class Profile(models.Model):
    class ChoicesSex(models.TextChoices):
        MAN = 'man', _('Мужской')
        WOMAN = 'woman', _('Женский')

    user = models.OneToOneField(User, on_delete=models.CASCADE)
    name = models.CharField(max_length=100, verbose_name='Имя')
    sex = models.CharField(
        max_length=5,
        choices=ChoicesSex.choices,
        default=ChoicesSex.MAN,
        verbose_name='Пол'
    )
    position = models.ForeignKey(Position, verbose_name='Должность', on_delete=models.SET_NULL, blank=True, null=True)
    half_time_work = models.BooleanField(verbose_name='Работа на пол ставки', default=False)
    avatar = models.ImageField(blank=True, null=True, verbose_name='Аватар', upload_to=generate_path_for_avatar)
    lab = models.ForeignKey(LabName, on_delete=models.CASCADE, verbose_name='Лаборатория', blank=True, null=True)
    room_number = models.ForeignKey(Room, on_delete=models.PROTECT, verbose_name='Номер кабинета',
                                    blank=True, null=True)
    number_of_elements = models.IntegerField(verbose_name='Количество позиций простых элементов', default=1, blank=True)
    robot = models.BooleanField(verbose_name='Робот или человек', default=False)
    can_change_big_object = models.BooleanField(
        verbose_name='Пользователь может изменять компоненты сложного объекта',
        default=False,
        blank=True,
    )
    tg_id = models.IntegerField(verbose_name='ID в телеграме', blank=True, null=True)
    tg_chat_id = models.IntegerField(verbose_name='ID телеграм чата', blank=True, null=True)
    tg_current_lab = models.ForeignKey(
        LabName, on_delete=models.CASCADE, verbose_name='Текущая лаборатория для телеграма', blank=True, null=True,
        related_name='telegram_lab_for_profile'
    )

    class Meta:
        ordering = ['name']
        verbose_name = 'Профиль сотрудника'
        verbose_name_plural = 'Профили сотрудников'

    def __str__(self):
        if self.lab is None:
            return self.name
        else:
            return f'{self.name} : {self.lab.name}'

    def save(self, *args, **kwargs):
        if not self.id:
            self.name = self.user.username
            if self.avatar:
                super().save(*args, **kwargs)
                self.save_avatar()
        else:
            old_self = Profile.objects.get(pk=self.pk)
            if old_self.avatar != self.avatar and self.avatar:
                old_self.delete_avatar()
                super().save(*args, **kwargs)
                self.save_avatar()
        super().save(*args, **kwargs)

    def get_short_name(self):
        """Возвращяет короткое имя пользователя"""
        name = self.name.split(' ')
        if len(name) == 3:
            return f'{name[0]} {name[1][0]}. {name[2][0]}.'
        elif len(name) == 2:
            return f'{name[0]} {name[1][0]}.'
        elif len(name) == 1:
            return f'{name[0]}'
        else:
            return 'Имя не задано'

    def get_absolute_url(self):
        return reverse('worker_page_url', kwargs={'pk': self.pk, 'lab': self.lab.slug})

    def avatar_url(self):
        if self.avatar:
            return self.avatar.url
        return '/media/avatars/default_avatar.png'

    def save_avatar(self):
        img = Image.open(self.avatar.path)
        img.save(self.avatar.path, quality=40)

    def get_or_create_month_for_calendar(self, year: int, month: int, admin=False):
        """
        month -> 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12
        """
        today = date.today()
        if today.year == year:
            if not abs(today.month - month) <= 2:
                return False
        elif abs(today.year - year) >= 1:
            if month not in [0, 1, 10, 11] or today.month not in [12, 1]:
                return False

        all_days = self.get_or_create_month(year, month)

        month_data = {}
        for day in all_days:
            if admin is True:
                href = day.get_absolute_url()
            else:
                href = None
            month_data[day.date.day] = {
                'href': href,
                'work_hours': day.work_hours,
                'work_minutes': day.work_minutes,
                'type': day.get_type_display(),
                'kod_type': day.type,
                'after_work_hours': day.after_work_hours,
            }
        month_data['status'] = 'ok'

        all_day_types = dict()
        for day_type in WorkCalendar.ChoicesDayType.choices:
            all_day_types[day_type[0]] = day_type[1]
        month_data['all_day_types'] = all_day_types

        return month_data

    def get_or_create_month(self, year, month):
        """
        month -> 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12
        """
        all_days = WorkCalendar.objects.filter(date__year=int(year), date__month=int(month), user=self)
        max_day = calendar.monthrange(int(year), int(month))[1]     # Количетсво дней в месяце
        if len(all_days) == 0:
            new_month = [self.create_day_for_user(year, month, day) for day in range(1, max_day + 1)]
            WorkCalendar.objects.bulk_create(new_month)
            return WorkCalendar.objects.filter(date__month=month, date__year=year, user=self)
        else:
            return all_days

    def create_day_for_user(self, year, month, day):
        new_day = WorkCalendar(date=date(year=int(year), month=int(month), day=day), user=self)
        str_day = new_day.date.strftime('%a').lower()
        if str_day == 'fri':
            if self.half_time_work is False:
                new_day.work_hours = '6'
                new_day.work_minutes = '00'
            else:
                new_day.work_hours = '3'
                new_day.work_minutes = '00'
        else:
            if str_day not in ('sat', 'sun'):
                if self.sex == 'man':
                    if self.half_time_work is False:
                        new_day.work_hours = '8'
                        new_day.work_minutes = '30'
                    else:
                        new_day.work_hours = '4'
                        new_day.work_minutes = '15'
                else:
                    if self.half_time_work is False:
                        new_day.work_hours = '7'
                        new_day.work_minutes = '30'
                    else:
                        new_day.work_hours = '3'
                        new_day.work_minutes = '45'
            else:
                new_day.work_hours = '0'
                new_day.work_minutes = '00'
                new_day.type = 'В'

        return new_day

    def delete_avatar(self):
        self.avatar.delete(save=False)


class WorkCalendar(models.Model):
    class ChoicesDayType(models.TextChoices):
        WORK = 'Ф', _('Рабочий день')
        DAY_OFF = 'В', _('Выходной')
        TELESCOPE = 'Т', _('Телескоп')
        VACATION = 'О', _('Отпуск')
        TRIP = 'К', _('Командировка')
        SICK = 'Б', _('Больничный')
        HOLIDAY_WORK = 'РП', _('Работа в праздничные дни')
        NIGHT_WORK = 'Н', _('Работа в ночное время')
        GOS_WORK = 'Г', _('Выполнение гособязанностей')
        BABY = 'Р', _('Отпуск в связи с родами')

    type = models.CharField(
        max_length=2,
        choices=ChoicesDayType.choices,
        default=ChoicesDayType.WORK,
        verbose_name='Тип дня'
    )
    work_hours = models.TextField(verbose_name='Количество рабочих часов', default='0')
    work_minutes = models.TextField(verbose_name='Количество рабочих минут в последнем часе', default='00')
    after_work_hours = models.FloatField(verbose_name='Часы сверхурочной работы', default=0)
    user = models.ForeignKey(to=Profile, on_delete=models.CASCADE, verbose_name='Профиль')
    date = models.DateField(verbose_name='Дата')

    class Meta:
        verbose_name = 'Календарь рабочих дней'
        verbose_name_plural = 'Календари рабочих дней'

    def __str__(self):
        return f'{self.user}, {self.date}, {self.type}'

    def get_absolute_url(self):
        return reverse('worker_calendar_change_date_form_url', kwargs={
            'pk': self.user.pk, 'lab': self.user.lab.slug, 'day': f'{self.date.day}-{self.date.month}-{self.date.year}'
        })

    def get_name_for_timesheet(self):
        if self.type == 'Ф' or self.type == 'Т':
            return {'type': 'Ф', 'hours': self.work_hours, 'minutes': self.work_minutes}
        else:
            return {'type': self.type}


class BaseObject(models.Model):
    class ChoicesStatus(models.TextChoices):
        NOT_IN_WORK = 'NW', _('Не в работе')
        IN_REPAIRING = 'IR', _('В ремонте')
        IN_WORK = 'IW', _('Активный')
        WRITTEN_OF = 'WO', _('Списано')

    name = models.CharField(max_length=200, verbose_name='Название')
    name_lower = models.CharField(max_length=200, verbose_name='Название в нижнем регистре', blank=True)
    slug = models.SlugField(max_length=250, unique=True, blank=True, verbose_name='URL')
    lab = models.ForeignKey(LabName, on_delete=models.CASCADE, verbose_name='Лаборатория')
    category = models.ForeignKey(Category, on_delete=models.CASCADE, verbose_name='Категория', blank=True, null=True)
    inventory_number = models.CharField(verbose_name='Инвентаризационный номер, не уникальный', max_length=100,
                                        blank=True, null=True)
    directory_code = models.CharField(verbose_name='Код справочника, не уникальный', max_length=100,
                                      blank=True, null=True)
    bill = models.CharField(max_length=100, verbose_name='Счет', blank=True, null=True)
    measure = models.CharField(verbose_name='Единица измерения', max_length=10, blank=True, null=True)
    total_price = models.FloatField(verbose_name='Сумма', default=0, blank=True)
    total_price_text = models.CharField(verbose_name='Общая сумма с пробелами', max_length=100, blank=True)
    amount = models.FloatField(verbose_name='Общее количество единиц', default=0)
    date_add = models.DateField(verbose_name='Дата принятия к учету', blank=True, null=True)
    status = models.CharField(
        max_length=2,
        choices=ChoicesStatus.choices,
        default=ChoicesStatus.IN_WORK,
        verbose_name='Статус'
    )

    class Meta:
        verbose_name = 'Базовый объект'
        verbose_name_plural = 'Базовые объекты'
        ordering = ['name_lower']

    def __str__(self):
        str_name = ''
        if self.inventory_number and self.bill:
            str_name += f'{self.inventory_number}, {self.bill}'
        elif self.inventory_number:
            str_name += f'{self.inventory_number}'
        elif self.bill:
            str_name += f'{self.bill}'
        return '{} ({})'.format(self.name, str_name)

    def save(self, *args, **kwargs):
        self.name = self.name.strip()
        if self.inventory_number:
            self.inventory_number = self.inventory_number.strip()
        if self.directory_code:
            self.directory_code = self.directory_code.strip()

        if self.pk is not None:
            old_self = BaseObject.objects.get(pk=self.pk)
            if old_self.name != self.name:
                self.name_lower = self.name.lower()
                self.create_slug()
            if old_self.lab != self.lab and self.category.lab != self.lab:
                new_category, created = Category.objects.get_or_create(
                    lab=self.lab, name=self.category.name, cat_type='BO'
                )
                self.category = new_category

            # if self.total_price != old_self.total_price or self.amount != old_self.amount:
            #     # Обновляем простой объект в том случае если у данного базового есть связь только с одним простым
            #     all_simple_objects = SimpleObject.objects.filter(base_object=self)
            #     if len(all_simple_objects) == 1:
            #         simple_object = all_simple_objects[0]
            #         simple_object.amount = self.amount
            #         simple_object.total_price = self.total_price
            #         simple_object.price = round(self.total_price / self.amount, 2)
            #         simple_object.save(update_base_object=False)
        else:
            self.name_lower = self.name.lower()
            self.create_slug()

        self.update_price()

        super().save(*args, **kwargs)

    def get_absolute_url(self):
        return reverse('base_object_page_url', kwargs={'slug': self.slug, 'lab': self.lab.slug})

    def get_absolute_update_url(self):
        return reverse('base_object_update_page_url', kwargs={'slug': self.slug, 'lab': self.lab.slug})

    def create_slug(self):
        all_slugs = BaseObject.objects.all().values_list('slug', flat=True)
        self.slug = gen_slug(lab=self.lab.slug, title=self.name, all_slugs=all_slugs)

    def update_price(self):
        """Обновление текстовой цены и общей стоимости"""
        if self.total_price != 0:
            self.total_price = round(self.total_price, 2)
            self.total_price_text = gen_text_price(self.total_price)
        else:
            self.total_price_text = '0,00'

    def get_short_bill(self):
        if self.bill:
            short_bill = self.bill.split('.')[1:3]
            if len(short_bill) > 0:
                return '.'.join(short_bill)
            else:
                return self.bill
        return ''

    def get_place(self):
        rooms = Room.objects.filter(pk__in=self.simpleobject_set.values_list('room'))
        return rooms

    def get_current_place(self):
        places = []
        for simple_obj in self.simpleobject_set.all():
            if simple_obj.place:
                places.append(simple_obj.place)
        return places


class InvoiceType(models.Model):
    invoice_type = models.CharField(max_length=200, verbose_name='Тип накладной')

    class Meta:
        verbose_name = 'Тип накладной'
        verbose_name_plural = 'Типы накладных'

    def __str__(self):
        return self.invoice_type


class Invoice(models.Model):
    invoice_type = models.ForeignKey(
        InvoiceType, verbose_name='Тип накладной', blank=True, null=True, on_delete=models.SET_NULL
    )
    number = models.CharField(max_length=50, verbose_name='Номер накладной')
    lab = models.ForeignKey(LabName, on_delete=models.CASCADE, related_name='invoices', related_query_name='invoice',
                            verbose_name='Лаборатория', blank=True, null=True)
    bill = models.CharField(max_length=100, verbose_name='Счет, субсчет', blank=True)
    date = models.DateField(verbose_name='Дата составления')
    total_price = models.FloatField(verbose_name='Сумма', default=0, blank=True)
    total_price_text = models.CharField(verbose_name='Общая сумма с пробелами', max_length=100, blank=True)

    class Meta:
        verbose_name = 'Накладная'
        verbose_name_plural = 'Накладные'
        ordering = ['-date']

    def __str__(self):
        return f'{self.lab}, {self.date}, {self.number}'

    def save(self, *args, **kwargs):
        self.update_price()
        super().save(*args, **kwargs)

    def update_price(self):
        """Обновление текстовой цены и общей стоимости"""
        if self.total_price != 0:
            self.total_price_text = gen_text_price(round(self.total_price, 2))
        else:
            self.total_price_text = '0,00'

    def get_absolute_url(self):
        return reverse('invoice_page_url', kwargs={'lab': self.lab.slug, 'pk': self.pk})

    def get_short_bill(self):
        if self.bill:
            short_bill = self.bill.split('.')[1:3]
            if len(short_bill) > 0:
                return '.'.join(short_bill)
            else:
                return self.bill
        return ''


class SimpleObject(models.Model):
    class ChoicesMeasure(models.TextChoices):
        GRAND = 'шт', _('шт')
        METRE = 'м', _('м')
        CENTIMETRE = 'см', _('см')
        KILOMETRE = 'км', _('км')
        BALLOON = 'бал', _('бал')
        KILOGRAM = 'кг', _('кг')
        GRAM = 'гр', _('гр')
        EMPTY = '---', _('---')

    base_object = models.ForeignKey(
        to=BaseObject, on_delete=models.CASCADE, blank=True, null=True, verbose_name='Базовая единица'
    )
    name = models.CharField(max_length=200, verbose_name='Название')
    name_lower = models.CharField(max_length=200, verbose_name='Название в нижнем регистре', blank=True)
    slug = models.SlugField(max_length=250, unique=True, blank=True, verbose_name='URL')
    lab = models.ForeignKey(LabName, on_delete=models.CASCADE, verbose_name='Лаборатория')
    category = models.ForeignKey(Category, on_delete=models.CASCADE, verbose_name='Категория', blank=True, null=True)
    room = models.ManyToManyField(Room, related_name='simple_object', related_query_name='simple_objects',
                                  verbose_name='Кабинет', blank=True)
    place = models.CharField(max_length=200, verbose_name='Место расположения', blank=True)
    price = models.FloatField(verbose_name='Стоимость одной единицы', default=0)
    total_price = models.FloatField(verbose_name='Сумма', default=0, blank=True)
    price_text = models.CharField(verbose_name='Сумма за единицу с пробелами', max_length=100, blank=True)
    total_price_text = models.CharField(verbose_name='Общая сумма с пробелами', max_length=100, blank=True)
    amount = models.FloatField(verbose_name='Общее количество единиц', default=0)
    amount_in_work = models.FloatField(verbose_name='Количество единиц в работе', blank=True, default=0)
    amount_free = models.FloatField(verbose_name='Количество свободных единиц', null=True, blank=True)
    text = models.TextField(blank=True, verbose_name='Описание')
    date_add = models.DateTimeField(default=timezone.now)
    measure = models.CharField(
        verbose_name='Единица измерения',
        choices=ChoicesMeasure.choices,
        default=ChoicesMeasure.EMPTY,
        max_length=3,
    )
    history = HistoricalRecords(
        verbose_name='История',
        excluded_fields=['name_lower', 'slug', 'lab', 'place', 'worker', 'price', 'total_price', 'total_price_text',
                         'amount_free', 'text', 'date_add']
    )   # История изменений

    class Meta:
        verbose_name = 'Простой объект'
        verbose_name_plural = 'Простые объекты'
        ordering = ['name_lower']

    def __str__(self):
        if self.base_object:
            if self.base_object.inventory_number and self.base_object.bill:
                return f'{self.name} ({self.base_object.inventory_number}, {self.base_object.bill})'
            elif self.base_object.inventory_number:
                return f'{self.name} ({self.base_object.inventory_number})'
            elif self.base_object.bill:
                return f'{self.name} ({self.base_object.bill})'
            else:
                return '{}'.format(self.name)
        return '{}'.format(self.name)

    def save(self, update_base_object=True, *args, **kwargs):
        update_big_objects_price = False    # Статус для обновления всех связанных объектов
        if self.pk is not None:
            old_self = SimpleObject.objects.get(pk=self.pk)
            if old_self.name != self.name:
                self.name = self.name.strip()
                self.create_slug()
                self.name_lower = self.name.lower()
            if old_self.price != self.price or old_self.amount != self.amount:
                """Если цена изменилась то считаем новую сумму и обновляем все связанные цены"""
                self.total_price = self.price * self.amount
                update_big_objects_price = True
                # self.update_price()

            if old_self.lab != self.lab and self.base_object:
                self.base_object.lab = self.lab

            self.update_amount()
            self.update_price()

            """Проверка изменений полей для записи истории изменений.
            Если какое-либо из полей изменилось то сохраняем объект с записями в истории,
            иначе пропускаем запись истории"""
            obj = [self.name, self.category, self.price_text,
                   self.amount, self.amount_in_work]
            old_obj = [old_self.name, old_self.category,
                       old_self.price_text, old_self.amount, old_self.amount_in_work]

            if obj == old_obj:
                self.skip_history_when_saving = True

        else:
            self.create_slug()
            self.total_price = self.price * self.amount
            self.name = self.name.strip()
            self.name_lower = self.name.lower()

            self.update_amount()
            self.update_price()

        if self.category is None:
            category, created = Category.objects.get_or_create(
                name='Без категории',
                lab=self.lab,
                cat_type='SO'
            )

            self.category = category

        super().save(*args, **kwargs)

        try:
            del self.skip_history_when_saving
        except:
            pass

        if update_big_objects_price:
            self.update_big_objects_price()

        if update_base_object:
            pass
            # self.update_base_object()

    def get_absolute_url(self):
        return reverse('simple_object_url', kwargs={'slug': self.slug, 'lab': self.lab.slug})

    def get_absolute_update_url(self):
        return reverse('simple_object_update_form_url', kwargs={'slug': self.slug, 'lab': self.lab.slug})

    def create_slug(self):
        all_slugs = SimpleObject.objects.all().values_list('slug', flat=True)
        self.slug = gen_slug(lab=self.lab.slug, title=self.name, all_slugs=all_slugs)

    def update_price(self):
        """Обновление текстовой цены и общей стоимости"""
        self.total_price = round(self.price * self.amount, 2)

        if self.price != 0:
            self.price_text = gen_text_price(round(self.price, 2))
            self.total_price_text = gen_text_price(round(self.total_price, 2))
        else:
            self.price_text = '0,00'
            self.total_price_text = '0,00'

        SimpleObject.objects.filter(pk=self.pk).update(
            total_price=self.total_price, price_text=self.price_text, total_price_text=self.total_price_text
        )

    def update_amount(self, update_big_objects_price=False):
        """Обновлние количества"""
        amount_in_work = 0

        self.amount = round(self.amount, 3)

        big_objects_list = BigObjectList.objects.filter(simple_object=self)
        for obj in big_objects_list:

            all_parts = BigObject.objects.filter(base__simple_components=obj, status='IW')
            for _ in all_parts:
                amount_in_work += obj.amount

        self.amount_in_work = amount_in_work
        self.amount_free = round(self.amount - self.amount_in_work, 2)

        SimpleObject.objects.filter(pk=self.pk).update(
            amount_free=self.amount_free, amount_in_work=self.amount_in_work, amount=self.amount
        )
        self.update_price()

        if update_big_objects_price:
            self.update_big_objects_price()

        self.update_base_object()

    def update_base_object(self):
        """Проверка на количество простых объектов входящих в состав базового.
        Если базовый объект состоит из одного простого, то их количество должно совпадать
        Если базовый объект состоит из нескольких простых, то его количество будет равно нулю только в случае если
        количество всех простых компонетов так же равно нулю"""
        if self.base_object:
            base_object_components = SimpleObject.objects.filter(base_object=self.base_object)
            if len(base_object_components) == 1:
                self.base_object.amount = self.amount
                self.base_object.total_price = self.total_price
                self.base_object.save()
            elif len(base_object_components) > 1:
                empty_components = 0
                total_price = 0
                for component in base_object_components:
                    total_price += component.total_price
                    if component.amount == 0:
                        empty_components += 1
                if empty_components != 0:
                    self.base_object.amount = 0
                    self.base_object.total_price = 0
                else:
                    self.base_object.amount = self.amount
                    self.base_object.total_price = total_price
                self.base_object.save()

    def update_big_objects_price(self):
        """Обновляем цену у всех комплектующих и больших объектов где присутствует данный простой объект"""
        base_big_objects = set()
        for component in BigObjectList.objects.filter(simple_object=self):
            """Обновление всех компонентов"""
            component.save()
            base_big_objects.add(component.big_object)

        for top_object in BigObject.objects.filter(status__in=['IW', 'NW'], top_level=True, base__in=base_big_objects):
            top_object.update_total_price()

    def write_off(self, amount):
        self.amount = self.amount - amount
        if self.amount == 0 and self.base_object:
            if len(self.base_object.simpleobject_set.all()) == 1:
                self.base_object.status = 'WO'
                self.base_object.amount = 0
                self.base_object.save()
        if self.amount == 0:
            for room in self.room.all():
                self.room.remove(room)
            WorkerEquipment.objects.filter(simple_object=self, order__isnull=True).delete()

        self.save()

    def delete(self, *args, **kwargs):
        all_big_objects = BigObject.objects.filter(base__simple_components__simple_object=self)
        for big_object in all_big_objects:
            simple_objects_list = BigObjectList.objects.filter(big_object=big_object.base).exclude(simple_object=self)
            big_object.update_total_price(simple_objects_list=simple_objects_list)
            BigObject.objects.filter(pk=big_object.pk).update(price=big_object.price, price_text=big_object.price_text)
        super().delete(*args, **kwargs)


class InvoiceBaseObject(models.Model):
    invoice = models.ForeignKey(Invoice, verbose_name='Накладная', on_delete=models.CASCADE)
    base_object = models.ForeignKey(BaseObject, verbose_name='Объект', on_delete=models.CASCADE)
    amount = models.IntegerField(verbose_name='Количество')

    def __str__(self):
        return f'Накладная {self.invoice.number} , объект {self.base_object.name} , количество {self.amount}'


class BaseBigObject(models.Model):
    name = models.CharField(max_length=200, verbose_name='Название')
    slug = models.SlugField(max_length=250, unique=True, blank=True, verbose_name='URL')
    text = models.TextField(verbose_name='Описание', blank=True)
    lab = models.ForeignKey(LabName, on_delete=models.CASCADE, verbose_name='Лаборатория', blank=True, null=True)
    category = models.ForeignKey(Category, on_delete=models.CASCADE, verbose_name='Категория', blank=True, null=True)
    inventory_number = models.CharField(verbose_name='Инвентаризационный номер', unique=True, blank=True, null=True,
                                        max_length=100)
    kod = models.CharField(verbose_name='РЮКС/РШАП', unique=True, blank=True, null=True, max_length=100)
    number_of_elements = models.IntegerField(verbose_name='Количество позиций простых элементов', default=1, blank=True)
    ready = models.BooleanField(verbose_name='Объект готов, запретить редактирование', default=False)

    class Meta:
        verbose_name = 'Основа сложного объекта'
        verbose_name_plural = 'Основы сложных объектов'
        ordering = ['name']

    def __str__(self):
        if self.kod:
            return '{}, {}'.format(self.name, self.kod)
        else:
            return self.name

    def save(self, *args, **kwargs):
        try:
            top_level = kwargs.pop('top_level')
        except KeyError:
            top_level = False

        create_big_object = False
        if self.pk is not None:
            old_self = BaseBigObject.objects.get(pk=self.pk)

            if old_self.name != self.name:
                self.create_slug()

            if old_self.inventory_number != self.inventory_number:
                self.inventory_number = self.inventory_number.upper()
        else:
            create_big_object = True
            self.create_slug()
            if self.inventory_number:
                self.inventory_number = self.inventory_number.upper()

        super().save(*args, **kwargs)

        if create_big_object:
            BigObject(
                base=self,
                top_level=top_level
            ).save()

    def create_slug(self):
        all_slugs = BaseBigObject.objects.all().values_list('slug', flat=True)
        self.slug = gen_slug(lab=self.lab.slug, title=self.name, all_slugs=all_slugs)

    def get_absolute_url(self):
        return reverse('base_big_object_page_url', kwargs={'slug': self.slug, 'lab': self.lab.slug})

    def get_unique_parts(self, include_self=None):
        """Находим сложный объект без родителей, который ссылается на данный базовый объект.
           Таким образом получаем основной экземпляр сложного объекта и находим всех его детей"""
        if include_self is None:
            include_self = True
        # all_parts = BigObject.objects.filter(base=self, full_name=self.name)
        all_parts = BigObject.objects.filter(base=self, parent=None)
        if all_parts:
            part = all_parts[0]
            return part.get_descendants(include_self=include_self)
        return None

    def get_top_level_big_objects(self):
        results = []
        components = set(self.components.all())
        all_top_level_objects = BigObject.objects.filter(top_level=True)
        for top_object in all_top_level_objects:
            all_children = set(top_object.get_descendants(include_self=True))
            if all_children.isdisjoint(components) is False:
                results.append(top_object)

        return results

    def get_base_big_object_parents(self):
        results = set()
        components = set(self.components.all())
        for component in components:
            if component.parent:
                if component.parent.base not in results:
                    results.add(component.parent.base)

        return results

    def update_price_for_instance(self):
        instance = BigObject.objects.filter(base=self, status__in=['IW', 'RD', 'NW'], parent__isnull=True)
        for big_object in instance:
            if not big_object.top_level:
                update_big_objects_price(self, update_child_price=True)
            else:
                big_object.update_total_price(update_child_price=False)


class BigObject(MPTTModel):
    class ChoicesStatus(models.TextChoices):
        NOT_IN_WORK = 'NW', _('Не в работе')
        IN_WORK = 'IW', _('В работе')
        READY = 'RD', _('Готово')
        WRITTEN_OF = 'WO', _('Списано')
    base = models.ForeignKey(to=BaseBigObject, on_delete=models.CASCADE, related_name='components',
                             related_query_name='component', verbose_name='Базовый объект',
                             blank=True, null=True)
    parent = TreeForeignKey('self', blank=True, null=True, related_name='children', verbose_name='Родитель',
                            on_delete=models.CASCADE)
    name = models.CharField(max_length=200, verbose_name='Название', blank=True, null=True)
    full_name = models.CharField(max_length=300, verbose_name='Полное название', blank=True, null=True)
    top_level = models.BooleanField(verbose_name='Объект верхнего уровня', default=False)
    status = models.CharField(
        max_length=2,
        choices=ChoicesStatus.choices,
        default=ChoicesStatus.NOT_IN_WORK,
        verbose_name='Статус'
    )
    kod_end = models.CharField(
        verbose_name='Окончание кода РЮКС/РШАП для конкретного экземпляра',
        blank=True, null=True, max_length=20
    )
    price = models.FloatField(verbose_name='Стоимость', default=0)
    price_text = models.CharField(verbose_name='Стоимость с пробелами', max_length=100, blank=True)
    system_number = models.CharField(max_length=400, verbose_name='Номер системы', blank=True, null=True)
    controller = models.CharField(max_length=100, verbose_name='Контроллер', blank=True, null=True)
    detector = models.CharField(max_length=400, verbose_name='Детектор', blank=True, null=True)
    interface = models.CharField(max_length=400, verbose_name='Интерфейс', blank=True, null=True)
    report = models.CharField(max_length=100, verbose_name='Отчет', blank=True, null=True)
    year = models.IntegerField(verbose_name='Год', blank=True, null=True)
    number_of_elements = models.IntegerField(verbose_name='Количество позиций простых элементов', default=1, blank=True)
    history = HistoricalRecords(
        verbose_name='История',
        excluded_fields=['slug', 'text', 'lab', 'price', 'number_of_elements',
                         'system_number', 'controller', 'detector', 'interface', 'report', 'year',
                         'level', 'lft', 'rght', 'tree_id']
    )  # История изменений

    class Meta:
        verbose_name = 'Сложный объект'
        verbose_name_plural = 'Сложные объекты'
        ordering = ['full_name', 'name']

    class MPTTMeta:
        order_insertion_by = ['full_name', 'name']

    def get_full_kod(self):
        if self.base.kod:
            if self.kod_end:
                full_kod = '{}-{}'.format(self.base.kod, self.kod_end)
            else:
                full_kod = self.base.kod
            return full_kod
        else:
            return ''

    def __str__(self):
        if self.name:
            if self.base.kod:
                full_path = ['{}, {}'.format(self.name, self.get_full_kod())]
            else:
                full_path = [self.name]
        else:
            if self.base.kod:
                full_path = ['{}, {}'.format(self.base.name, self.get_full_kod())]
            else:
                full_path = [self.base.name]

        k = self.parent
        while k is not None:
            # full_path.append(k.base.name)
            full_path.append('---')

            k = k.parent
        # return ' -> '.join(full_path[::-1])
        return ''.join(full_path[::-1])

    def save(self, *args, **kwargs):
        full_path = [self.base.name]
        k = self.parent
        while k is not None:
            full_path.append(k.base.name)
            k = k.parent
        self.full_name = ' -> '.join(full_path[::-1])

        if self.pk is not None:
            old_self = BigObject.objects.get(pk=self.pk)

            if self.status != old_self.status:
                all_children = self.get_descendants(include_self=True)
                all_children.update(status=self.status)
                for child in all_children:
                    child.update_simple_objects(old_self=old_self)

            """Проверка изменений полей для записи истории изменений.
            Если какое-либо из полей изменилось то сохраняем объект с записями в истории,
            иначе пропускаем запись истории"""
            obj = [self.base.name, self.status, self.base.inventory_number, self.price_text]
            old_obj = [old_self.base.name, old_self.status, old_self.base.inventory_number, old_self.price_text]
            if obj == old_obj:
                self.skip_history_when_saving = True

            super().save(*args, **kwargs)
            try:
                del self.skip_history_when_saving
            except:
                pass
            # self.update_simple_objects(old_self=old_self)  # Обновление простых объектов входящих в состав сложного

        else:
            if self.status in ['WO', 'NW']:
                """Если статус не рабочий тогда зануляем стоимость"""
                self.price = 0
                self.price_text = '0,00'
            super().save(*args, **kwargs)

        if self.status in ['NW', 'IW', 'RD']:
            self.update_total_price()

            # BigObject.objects.rebuild()

    def get_absolute_url(self):
        return reverse(
            'big_object_page_url', kwargs={
                'slug': self.base.slug, 'lab': self.base.lab.slug, 'pk': self.pk
            }
        )

    def copy_object_and_children(self, new_name=None, kod_end=None):
        first_children = self.get_children()
        new_top_part = self
        new_top_part.pk = None
        new_top_part.status = 'NW'
        # new_top_part.top_level = False
        if new_name:
            new_top_part.name = new_name
        if kod_end:
            new_top_part.kod_end = kod_end
        new_top_part.save()

        def search_all_parts(first_part):
            for child in first_children:
                def create_new_part(p, new_parent):
                    p_children = p.get_children()

                    p_new = p
                    p_new.pk = None
                    p_new.parent = new_parent
                    p_new.status = 'NW'
                    p_new.top_level = False
                    p_new.save()

                    for i in p_children:
                        create_new_part(i, p_new)

                create_new_part(child, first_part)

        search_all_parts(new_top_part)

        BigObject.objects.rebuild()

        return new_top_part

    def copy_children(self):
        pass

    def update_total_price(self, simple_objects_list=None, update_child_price=True):
        """Обновление стоимости для сложного объекта с учетом стоимости всех детей"""
        if simple_objects_list is None:
            simple_objects_list = BigObjectList.objects.filter(big_object=self.base)
        total_price = 0
        for simple_object in simple_objects_list:
            total_price += simple_object.total_price
        for child in self.get_descendants(include_self=False)[::-1]:
            if update_child_price:
                child.update_self_price()
            total_price += child.price
        self.price = total_price
        if self.price == 0:
            self.price_text = '0,00'
        else:
            self.price_text = gen_text_price(self.price)
        BigObject.objects.filter(pk=self.pk).update(price=self.price, price_text=self.price_text)

    def update_self_price(self):
        """Обновление стоимости конкретно для данного сложного объекта без учета детей"""
        price = 0
        simple_objects_list = BigObjectList.objects.filter(big_object=self.base)
        for simple_object in simple_objects_list:
            price += simple_object.total_price
        self.price = price
        if self.price == 0:
            self.price_text = '0,00'
        else:
            self.price_text = gen_text_price(self.price)

        BigObject.objects.filter(pk=self.pk).update(price=self.price, price_text=self.price_text)

    def update_simple_objects(self, old_self):
        if self.status == 'IW':
            """Если статус рабочий, тогда обновляем количество свободных простых объектов"""
            # self.update_price()
            all_simple_objects = SimpleObject.objects.filter(big_object_list__big_object=self.base)
            for simple_object in all_simple_objects:
                simple_object.update_amount()

        elif self.status in ['WO', 'NW', 'RD']:
            """Если старый статус был в работе а новый нет, тогда зануляем цену и обновляем свободные простые объекты"""
            all_simple_objects = SimpleObject.objects.filter(big_object_list__big_object=self.base)
            for simple_object in all_simple_objects:
                if self.status == 'RD' and old_self.status == 'IW':
                    """Если статус изменился с рабочего на готовый, тогда вычитаем количество использованных
                    простых объектов в этой камере из общего числа этих объектов"""
                    amount = BigObjectList.objects.get(big_object=self.base, simple_object=simple_object).amount
                    simple_object.amount -= amount
                simple_object.update_amount()
            if self.status == 'WO' and old_self.status == 'RD':
                """После сохранения всех простых объектов списываем их если статус
                сложного объекта изменился на 'списано' и количество простых объектов == 0"""
                self.write_off()
                for simple_object in all_simple_objects.filter(room__isnull=False, amount=0):
                    for room in simple_object.room.all():
                        simple_object.room.remove(room)
                    simple_object.save()
                    WorkerEquipment.objects.filter(simple_object=simple_object, order__isnull=True).delete()

    def write_off(self):
        """Списание сложного объекта"""
        all_simple_objects_for_self = SimpleObject.objects.filter(
            big_object_list__big_object=self.base
        )   # Список всех простых объектов входящих в состав сложного
        for simple_object in all_simple_objects_for_self:
            simple_object.update_amount()   # Обновляем количество и стоимость
            if simple_object.amount == 0:   # Если количество простых равно нулю тогда находим их базовые
                base = simple_object.base_object
                if len(base.simpleobject_set.all()) == 1:   # Если у базового только один простой, списываем
                    base.status = 'WO'
                    base.save()

    def delete(self, *args, **kwargs):
        self.status = 'WO'
        self.save()
        super().delete(*args, **kwargs)


class FileAndImageCategory(models.Model):
    simple_object = models.ForeignKey(
        SimpleObject, blank=True, null=True, on_delete=models.CASCADE, verbose_name='Простой объект'
    )
    big_object = models.ForeignKey(
        BaseBigObject, blank=True, null=True, on_delete=models.CASCADE, verbose_name='Сложный объект'
    )
    invoice = models.ForeignKey(
        Invoice, blank=True, null=True, on_delete=models.CASCADE, verbose_name='Накладная'
    )
    name = models.CharField(max_length=150, verbose_name='Название категории')
    text = models.TextField(max_length=2000, verbose_name='Описание', blank=True)
    slug = models.SlugField(max_length=250, blank=True, null=True, verbose_name='Название на латинице')

    class Meta:
        verbose_name = 'Категория документов'
        verbose_name_plural = 'Категории документов'

    def __str__(self):
        if self.simple_object:
            return '{} , для {}'.format(self.name, self.simple_object.name)
        elif self.big_object:
            return '{} , для {}'.format(self.name, self.big_object.name)
        elif self.invoice:
            return '{} , для накладной № {}'.format(self.name, self.invoice.number)
        else:
            return '{}'.format(self.name)

    def save(self, *args, **kwargs):
        if self.pk is None:
            self.slug = gen_slug_for_categories(self.name)
        else:
            old_self = FileAndImageCategory.objects.get(pk=self.pk)
            if old_self.name != self.name:
                self.slug = gen_slug_for_categories(self.name)
        super().save(*args, **kwargs)


class ImageForObject(models.Model):
    image = models.ImageField(verbose_name='Изображение', upload_to=generate_path)
    image_big = models.ImageField(blank=True, upload_to=generate_path,
                                  verbose_name='Фото без сжатия, создается само при сохранеии')
    category = models.ForeignKey(
        FileAndImageCategory, blank=True, null=True, on_delete=models.CASCADE, verbose_name='Категория',
        related_name='images', related_query_name='image'
    )

    class Meta:
        verbose_name = 'Изображение'
        verbose_name_plural = 'Изображения для объектов'

    def __str__(self):
        if self.category.simple_object:
            return 'Изображение для ' + str(self.category.simple_object.name)
        elif self.category.big_object:
            return 'Изображение для ' + str(self.category.big_object.name)
        else:
            return 'Изображение ' + str(self.image.name)

    def filename(self):
        return os.path.basename(self.image.name)

    def save(self, *args, **kwargs):
        if self.pk is not None:
            old_self = ImageForObject.objects.get(pk=self.pk)
            if self.image != old_self.image:
                if len(self.image.name) > 50:
                    self.image.name = str(self.image.name)[50::]

                super().save(*args, **kwargs)
                some_model_save_photo(self, 700, 700)
        else:
            if len(self.image.name) > 50:
                self.image.name = str(self.image.name)[50::]

            super().save(*args, **kwargs)
            some_model_save_photo(self, 700, 700)

        super().save(*args, **kwargs)

    def change_photo(self, new_image):
        self.image.delete(False)
        self.image_big.delete(False)
        self.image = new_image
        self.save()

    def change_path(self, *args, **kwargs):
        for image in [self.image, self.image_big]:
            image_name = str(image.path).rsplit(os.sep, 1)[1]
            image.name = generate_path(self, image_name)
        super().save(*args, **kwargs)

    def delete(self, *args, **kwargs):
        full_path = str(self.image.path).rsplit(os.sep, 1)[0]
        self.image.delete(False)
        self.image_big.delete(False)
        del_path(full_path)
        super().delete(*args, **kwargs)


class FileForObject(models.Model):
    file = models.FileField(upload_to=generate_path_for_files, verbose_name='Файл')
    category = models.ForeignKey(
        FileAndImageCategory, blank=True, null=True, on_delete=models.CASCADE, verbose_name='Категория',
        related_name='files', related_query_name='file'
    )

    class Meta:
        verbose_name = 'Файл'
        verbose_name_plural = 'Файлы для объектов'

    def __str__(self):
        if self.category.simple_object:
            return 'Файл для ' + str(self.category.simple_object.name)
        elif self.category.big_object:
            return 'Файл для ' + str(self.category.big_object.name)
        else:
            return 'Файл ' + str(self.file.name)

    def filename(self):
        return os.path.basename(self.file.name)

    def delete(self, *args, **kwargs):
        full_path = str(self.file.path).rsplit(os.sep, 1)[0]
        self.file.delete(save=False)
        del_path(full_path)
        super().delete(*args, **kwargs)


class BigObjectList(models.Model):
    simple_object = models.ForeignKey(
        to=SimpleObject,
        on_delete=models.CASCADE,
        related_name='big_objects_list',
        related_query_name='big_object_list',
        verbose_name='Простой объект',
    )
    amount = models.FloatField(verbose_name='Количество', default=0)
    total_price = models.FloatField(verbose_name='Сумма', default=0, blank=True)
    total_price_text = models.CharField(verbose_name='Сумма с пробелами', max_length=100, blank=True)
    big_object = models.ForeignKey(to=BaseBigObject, verbose_name='Сложный объект', on_delete=models.CASCADE,
                                   blank=True, related_name='simple_components')
    history = HistoricalRecords(verbose_name='История')  # История изменений

    class Meta:
        verbose_name = 'Составляющая сложного объекта'
        verbose_name_plural = 'Составляющие сложного объекта'
        ordering = ['simple_object__name_lower']

    def __str__(self):
        return f'{self.simple_object.name}, количество : {self.amount}'

    def save(self, update_simple_object=False, *args, **kwargs):
        self.update_total_price()
        super().save(*args, **kwargs)
        if update_simple_object:
            self.simple_object.update_amount()

    def search_all_top_level_objects_with_simple_object(self):
        result = []
        for big_object in BigObject.objects.filter(top_level=True, status='IW'):
            for child in big_object.get_descendants(include_self=True):
                component = child.base.simple_components.filter(simple_object=self.simple_object)
                if component:
                    result.append(big_object)
                    break
        if len(result) > 0:
            return result
        else:
            return False

    def update_total_price(self):
        self.total_price = self.amount * self.simple_object.price
        self.total_price_text = gen_text_price(self.total_price)

    def delete(self, *args, **kwargs):
        base_big_object = self.big_object
        simple_object = self.simple_object
        super().delete()
        simple_object.update_amount(update_big_objects_price=False)
        update_big_objects_price(base_big_object)


class Order(models.Model):
    date = models.DateTimeField(verbose_name='Дата запроса', default=timezone.now)
    number_of_elements = models.IntegerField(verbose_name='Количество позиций простых элементов', default=1, blank=True)
    confirm = models.BooleanField(verbose_name='Подтверждение', default=False)
    lab = models.ForeignKey(LabName, on_delete=models.CASCADE, related_name='order', related_query_name='orders',
                            verbose_name='Лаборатория', blank=True, null=True)

    class Meta:
        verbose_name = 'Заявка на выдачу'
        verbose_name_plural = 'Заявки на выдачу'
        ordering = ['-date']

    def __str__(self):
        return f'Заявка : {self.date}'

    def save(self, *args, **kwargs):
        if self.pk is not None:
            old_order = Order.objects.get(pk=self.pk)
            if old_order.confirm is False and self.confirm is True:
                self.update_worker_equipment()
        super().save(*args, **kwargs)

    def update_worker_equipment(self):
        order_equipments = WorkerEquipment.objects.filter(order=self)
        for eq in order_equipments:
            equipment, created = WorkerEquipment.objects.get_or_create(
                profile=eq.profile, order__isnull=True, simple_object=eq.simple_object
            )
            if created:
                equipment.amount = eq.amount
            else:
                equipment.amount += eq.amount

            equipment.save()


class WorkerEquipment(models.Model):
    profile = models.ForeignKey(to=Profile, on_delete=models.CASCADE, verbose_name='Профиль', blank=True)
    simple_object = models.ForeignKey(
        to=SimpleObject,
        on_delete=models.CASCADE,
        related_name='worker_objects_list',
        related_query_name='worker_object_list',
        verbose_name='Объект'
    )
    amount = models.FloatField(verbose_name='Количество', default=0)
    price = models.FloatField(verbose_name='Цена одной позиции', blank=True, null=True)
    total_price = models.FloatField(verbose_name='Сумма данной позиции', blank=True, null=True)
    order = models.ForeignKey(Order, on_delete=models.CASCADE, related_name='equipment',
                              related_query_name='equipments', verbose_name='Заказ', blank=True, null=True)

    class Meta:
        verbose_name = 'Оборудование у сотрудника'
        verbose_name_plural = 'Обородувания у сотрудников'
        ordering = ['simple_object__name_lower']

    def __str__(self):
        return f'{self.simple_object.name}, количетво : {self.amount}'

    def save(self, *args, **kwargs):
        if self.simple_object:
            if self.simple_object.price and self.simple_object.amount:
                self.price = self.simple_object.price
                self.total_price = round(self.price * self.amount, 2)
        super().save(*args, **kwargs)


class DataBaseDoc(models.Model):
    name = models.CharField(max_length=150, verbose_name='Название файла')
    lab = models.ForeignKey(LabName, on_delete=models.PROTECT, verbose_name='Лаборатория', blank=True, null=True)
    file = models.FileField(upload_to=generate_path_for_database, verbose_name='Файл')
    date_add = models.DateTimeField(default=timezone.now)

    class Meta:
        verbose_name = 'Файл с данными'
        verbose_name_plural = 'Файлы с данными'

    def __str__(self):
        if self.lab:
            return '{} {}'.format(self.name, self.lab.name)
        else:
            return self.name

    def update_all_data(self, create_simple_objects=False):
        from openpyxl import load_workbook
        wb = load_workbook(self.file.path)
        sheet = wb.worksheets[0]
        for row in sheet.iter_rows(values_only=True, min_row=1):
            category, created = Category.objects.get_or_create(name=row[3], lab=self.lab, cat_type='BO')

            base_object = BaseObject(
                name=row[2].strip(),
                lab=self.lab,
                inventory_number=row[0],
                directory_code=row[1],
                measure=row[5],
                total_price=float(row[7]),
                amount=float(row[6]),
                category=category,
            )
            base_object.save()

            if create_simple_objects is True:
                if row[4] is not None:
                    simple_category, created = Category.objects.get_or_create(name=row[4], lab=self.lab, cat_type='SO')
                else:
                    simple_category, created = Category.objects.get_or_create(
                        name='Разное ({})'.format(category.name.lower()), lab=self.lab, cat_type='SO'
                    )

                SimpleObject(
                    base_object=base_object,
                    name=base_object.name,
                    lab=self.lab,
                    measure=base_object.measure,
                    price=round(base_object.total_price / base_object.amount, 2),
                    amount=base_object.amount,
                    category=simple_category,
                ).save()

    def delete(self, *args, **kwargs):
        self.file.delete(save=False)
        super().delete(*args, **kwargs)


@receiver(post_save, sender=User)
def create_user_profile(sender, instance, created, **kwargs):
    if created:
        Profile.objects.create(user=instance)


@receiver(post_save, sender=User)
def save_user_profile(sender, instance, **kwargs):
    instance.profile.save()
